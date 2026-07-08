"""Locks in the amortization-workbook parser for the Loan Documents builder.

Synthetic workbooks mirror the two real layouts:
- "Balloon": fee block in G/H starting "Gross Loan Amount", ending
  "To be disbursed to Borrower (Est)"; Sheet1 schedule headed at C7.
- "Fully Amortized": fee block in K/L (with a blank row inside), the main
  sheet also carrying a full amortization grid whose header includes
  "Beginning Balance" (which must NOT be mistaken for the Exhibit A table).
"""

import io
from datetime import datetime

from openpyxl import Workbook

from app.loandocs_sheet import parse_settlement_workbook


def _bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _balloon_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Balloon"
    ws["B5"], ws["D5"] = "Loan Amount:", 5_400_000
    ws["G5"], ws["H5"] = "Gross Loan Amount", 5_400_000
    ws["G6"], ws["H6"] = "SRC Origination Fee", 216_000
    ws["G7"], ws["H7"] = "Referral Fee", 56_000
    ws["G8"], ws["H8"] = "Prior Lender Payoff", 2_300_000
    ws["G9"], ws["H9"] = "To be disbursed to Borrower (Est)", 2_828_000
    # full amortization grid — must not be picked up as the Exhibit A table
    ws["A18"], ws["B18"], ws["C18"] = "Payment Number", "Payment \nDate", "Beginning \nBalance"
    ws["G18"], ws["H18"] = "Principal", "Interest"
    ws["A19"], ws["B19"], ws["C19"] = 1, datetime(2024, 11, 1), 5_400_000
    ws["G19"], ws["H19"] = 0, 24_300

    s1 = wb.create_sheet("Sheet1")
    s1["C7"], s1["D7"], s1["E7"], s1["F7"], s1["G7"] = (
        "Payment Number", "Payment \nDate", "Principal", "Interest", "Total")
    s1["C8"], s1["D8"], s1["E8"], s1["F8"], s1["G8"] = 1, datetime(2024, 11, 1), 0, 0, 0
    s1["C9"], s1["D9"], s1["E9"], s1["F9"], s1["G9"] = (
        2, datetime(2025, 5, 15), 5_400_000, 228_150, 5_628_150)
    # totals row: no payment number -> table ends here
    s1["E10"], s1["F10"], s1["G10"] = 5_400_000, 228_150, 5_628_150
    return _bytes(wb)


def _fully_amortized_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Fully Amortized"
    ws["K4"], ws["L4"] = "Gross Loan Amount", 1_000_000
    # blank K5 row inside the block — the walk must continue past it
    ws["K6"], ws["L6"] = "SSL underwriting Fee", 10_000
    ws["K7"], ws["L7"] = "DDD Insurance", 58_000
    ws["K10"], ws["L10"] = "To be disbursed to Borrower (Est)", 932_000

    s1 = wb.create_sheet("Sheet1")
    s1["D8"], s1["E8"], s1["F8"], s1["G8"], s1["H8"] = (
        "Payment \nNumber", "Payment \nDate", "Principal", "Interest", "Total Payment")
    s1["D9"], s1["E9"], s1["F9"], s1["G9"], s1["H9"] = (
        1, datetime(2025, 1, 15), 27358.637533, 12916.666666, 40275.3042)
    s1["D10"], s1["E10"], s1["F10"], s1["G10"], s1["H10"] = (
        2, datetime(2025, 2, 15), 27712.019934, 12563.284265, 40275.3042)
    return _bytes(wb)


def test_balloon_layout():
    r = parse_settlement_workbook(_balloon_workbook())
    assert r["settlement_sheet"] == "Balloon"
    assert r["gross_loan_amount"] == 5_400_000
    assert [(l["label"], l["amount"]) for l in r["lines"]] == [
        ("SRC Origination Fee", 216_000),
        ("Referral Fee", 56_000),
        ("Prior Lender Payoff", 2_300_000),
    ]
    assert r["disbursed_check"] == 2_828_000
    assert r["schedule_sheet"] == "Sheet1"
    assert len(r["schedule"]) == 2  # totals row excluded
    assert r["schedule"][0] == {"date": "11/1/24", "interest": 0, "principal": 0, "total": 0}
    assert r["schedule"][1]["principal"] == 5_400_000


def test_fully_amortized_layout_with_blank_row():
    r = parse_settlement_workbook(_fully_amortized_workbook())
    assert r["gross_loan_amount"] == 1_000_000
    assert [l["label"] for l in r["lines"]] == ["SSL underwriting Fee", "DDD Insurance"]
    assert r["disbursed_check"] == 932_000
    assert len(r["schedule"]) == 2
    assert r["schedule"][0]["date"] == "1/15/25"
    assert r["schedule"][0]["interest"] == 12916.67  # rounded to cents
    assert r["schedule"][0]["total"] == 40275.3


def test_unrecognized_workbook_raises():
    wb = Workbook()
    wb.active["A1"] = "nothing to see here"
    try:
        parse_settlement_workbook(_bytes(wb))
        assert False, "expected ValueError"
    except ValueError as exc:
        assert "Gross Loan Amount" in str(exc)
