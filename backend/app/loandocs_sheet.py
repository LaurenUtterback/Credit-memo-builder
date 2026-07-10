"""Parse the deal's loan-amortization workbook for the Loan Documents builder.

The underwriting workbook ("Balloon *.xlsx" / "Fully Amortized *.xlsx") carries
two things the closing package needs:

1. The Memo of Settlement fee block — on the sheet named "Balloon" or
   "Fully Amortized" (found generically): a "Gross Loan Amount" label with the
   amount in a nearby column, the fee/payoff lines below it (SRC Origination
   Fee, insurance, payoffs, ...), then "To be disbursed to Borrower (Est)".
   Some workbooks continue past that subtotal with lines carved out of the
   to-Borrower figure (e.g. DDD Insurance) down to a final "Net to be
   disbursed to Borrower (Est)" — those come back as ``post_lines`` (kept only
   when the Net row is actually there, so a missing terminator can't sweep up
   unrelated cells below the block). The block's column position varies by
   workbook (G/H on Balloon, K/L on Fully Amortized), so it is located by the
   label text, never by coordinates.

2. The Note's Exhibit A repayment schedule — on "Sheet1": a header row of
   Payment Number | Payment Date | Principal | Interest | Total, one row per
   payment, then a totals row with no payment number (skipped; the template
   recomputes totals). The full amortization grid on the main sheet has a
   similar header but also "Beginning Balance", which is how it is excluded.

Everything is read from cached values (data_only=True) and reflected verbatim;
the sheet's "To be disbursed" / "Net to be disbursed" figures are returned only
as cross-checks — the rendered memo always recomputes both from the lines.
"""

from __future__ import annotations

import io
import re
from datetime import date, datetime

from openpyxl import load_workbook

_MAX_AMOUNT_OFFSET = 4      # amount sits within this many columns right of a label
_MAX_BLOCK_ROWS = 40        # fee block never runs longer than this


def _norm(v) -> str:
    return re.sub(r"\s+", " ", str(v)).strip().lower() if v is not None else ""


def _num(v):
    if isinstance(v, bool) or v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace("$", "").replace(",", "").strip())
    except ValueError:
        return None


def _fmt_date(v) -> str:
    if isinstance(v, (datetime, date)):
        return f"{v.month}/{v.day}/{str(v.year)[2:]}"
    return str(v).strip() if v is not None else ""


def _amount_right_of(ws, row: int, col: int):
    """First numeric value within a few columns right of (row, col)."""
    for c in range(col + 1, col + 1 + _MAX_AMOUNT_OFFSET):
        n = _num(ws.cell(row=row, column=c).value)
        if n is not None:
            return n, c
    return None, None


def _parse_fee_block(ws) -> dict | None:
    """Find 'Gross Loan Amount' and read the label/amount pairs below it."""
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 60)):
        for cell in row:
            if _norm(cell.value) != "gross loan amount":
                continue
            gross, amt_col = _amount_right_of(ws, cell.row, cell.column)
            if gross is None:
                continue
            lines, post_lines, disbursed, net = [], [], None, None
            after_disbursed = False
            for r in range(cell.row + 1, cell.row + 1 + _MAX_BLOCK_ROWS):
                label = ws.cell(row=r, column=cell.column).value
                if _norm(label) == "":
                    continue
                amount = _num(ws.cell(row=r, column=amt_col).value)
                if amount is None:
                    amount, _ = _amount_right_of(ws, r, cell.column)
                if "net to be disbursed" in _norm(label):
                    net = amount
                    break
                if "to be disbursed" in _norm(label):
                    disbursed = amount
                    after_disbursed = True
                    continue
                if amount is None:
                    continue
                (post_lines if after_disbursed else lines).append(
                    {"label": str(label).strip(), "amount": round(amount, 2)})
            if net is None:
                # no "Net to be disbursed" terminator: anything gathered after
                # the disbursed row is just cells below the block — drop it.
                post_lines = []
            return {
                "settlement_sheet": ws.title,
                "gross_loan_amount": round(gross, 2),
                "lines": lines,
                "disbursed_check": round(disbursed, 2) if disbursed is not None else None,
                "post_lines": post_lines,
                "net_check": round(net, 2) if net is not None else None,
            }
    return None


def _find_schedule_header(ws):
    """Header row + column map for the Exhibit A table, or None.

    Skips the workbook's full amortization grid, whose header also carries
    'Beginning Balance'.
    """
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 40)):
        labels = {_norm(c.value): c.column for c in row if c.value is not None}
        if not labels:
            continue
        if "beginning balance" in labels:
            continue
        cols = {}
        for text, col in labels.items():
            if text == "payment number":
                cols["num"] = col
            elif text == "payment date":
                cols["date"] = col
            elif text == "principal":
                cols["principal"] = col
            elif text == "interest":
                cols["interest"] = col
            elif text in ("total", "total payment"):
                cols["total"] = col
        if {"num", "date", "principal", "interest"} <= cols.keys():
            return row[0].row, cols
    return None


def _parse_schedule(wb) -> dict | None:
    # prefer the plainly named "Sheet1"-style tabs the template uses
    sheets = sorted(wb.worksheets,
                    key=lambda ws: 0 if _norm(ws.title).startswith("sheet") else 1)
    for ws in sheets:
        found = _find_schedule_header(ws)
        if not found:
            continue
        header_row, cols = found
        rows = []
        for r in range(header_row + 1, ws.max_row + 1):
            num = _num(ws.cell(row=r, column=cols["num"]).value)
            if num is None:
                break  # totals row (or end of table)
            interest = _num(ws.cell(row=r, column=cols["interest"]).value) or 0
            principal = _num(ws.cell(row=r, column=cols["principal"]).value) or 0
            total = None
            if "total" in cols:
                total = _num(ws.cell(row=r, column=cols["total"]).value)
            rows.append({
                "date": _fmt_date(ws.cell(row=r, column=cols["date"]).value),
                "interest": round(interest, 2),
                "principal": round(principal, 2),
                "total": round(total, 2) if total is not None
                         else round(interest + principal, 2),
            })
        if rows:
            return {"schedule_sheet": ws.title, "schedule": rows}
    return None


def parse_settlement_workbook(data: bytes) -> dict:
    """Bytes of an .xlsx -> settlement lines + Exhibit A schedule."""
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True)
    except Exception as exc:  # noqa: BLE001 - corrupt/foreign file
        raise ValueError(f"Could not open the workbook: {exc}") from exc

    fee = None
    for ws in wb.worksheets:
        fee = _parse_fee_block(ws)
        if fee:
            break
    sched = _parse_schedule(wb)

    if not fee and not sched:
        raise ValueError(
            'Found neither a "Gross Loan Amount" fee block nor a '
            '"Payment Number / Payment Date / Principal / Interest" schedule '
            "in this workbook.")

    result = {"settlement_sheet": "", "gross_loan_amount": None, "lines": [],
              "disbursed_check": None, "post_lines": [], "net_check": None,
              "schedule_sheet": "", "schedule": []}
    if fee:
        result.update(fee)
    if sched:
        result.update(sched)
    return result
