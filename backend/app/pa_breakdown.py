"""Parse a "Participant Breakdown" .xlsx into deal info + per-participant terms.

The breakdown sheet holds, on its data tab (named e.g. "Sheet1" or "$1.35M"):
  * a "Name" cell (borrower) and "Loan amount" and "Loan #" cells, and
  * a table with header row: Participant | Participant $ | Participant % |
    Origination %/Points % | Origination $/Points $ | Interest Rate |
    Share of Late fee | …
and, on another tab, a Participant -> Email lookup.

We reflect EXACTLY what the sheet shows: each percentage is read from its own
cell and rendered at the precision of that cell's number format (so a cell shown
as "18.52%" comes back as "18.52%", not a recomputed 18.5185%). When a derived
cell has no cached value (a freshly-edited file Excel never recalculated), we
fall back to the sheet's own formula — Participation % = Participant $ / Loan
amount, Origination $ = Origination % * Participant $, Late-fee share =
Participation % * 50% — still rendered at the cell's displayed precision.
"""

from __future__ import annotations

import io
import re

from openpyxl import load_workbook


def _norm(v) -> str:
    return " ".join(str(v).strip().lower().split()) if v is not None else ""


def _num(v) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _to_fraction(v) -> float:
    """A rate cell stores 0.14 for 14%; tolerate a literal 14 too."""
    n = _num(v)
    return n if -1.0 <= n <= 1.0 else n / 100.0


def _pct_decimals(fmt: str, default: int = 2) -> int:
    """Decimal places a percent number-format shows: '0.00%' -> 2, '0%' -> 0."""
    if not fmt or "%" not in fmt:
        return default
    m = re.search(r"\.(0+)\s*%", fmt)
    return len(m.group(1)) if m else 0


def _money(v) -> str:
    return f"${_num(v):,.2f}"


def _value_right_of(ws, label: str):
    target = _norm(label)
    for row in ws.iter_rows():
        for c in row:
            if _norm(c.value) == target:
                return ws.cell(c.row, c.column + 1).value
    return None


def _find_header(ws):
    """Return (header_row_index, {normalized_header: column_index}) or (None, {})."""
    for row in ws.iter_rows():
        norms = {_norm(c.value): c.column for c in row if c.value is not None}
        if "participant" in norms and "participant $" in norms:
            return row[0].row, norms
    return None, {}


def _find_data_sheet(wb):
    for ws in wb.worksheets:
        row, cols = _find_header(ws)
        if row:
            return ws, row, cols
    return None, None, {}


def _email_map(wb) -> dict:
    """Scan every sheet for rows containing an email; key by normalized name."""
    out = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            cells = [c.value for c in row if c.value is not None]
            email = next((str(c) for c in cells if "@" in str(c)), None)
            if not email:
                continue
            name = next((str(c) for c in cells if "@" not in str(c)
                         and _norm(c) not in ("email", "paricipant", "participant")), None)
            if name:
                out[_norm(name)] = email.strip()
    return out


def parse_breakdown(data: bytes) -> dict:
    """Parse breakdown bytes -> {deal:{...}, participants:[{...}]}.

    Percentage and dollar fields come back as display-ready strings that match
    the sheet. Raises ValueError if the sheet isn't a recognizable breakdown.
    """
    wb = load_workbook(io.BytesIO(data), data_only=True)
    ws, hdr_row, cols = _find_data_sheet(wb)
    if not ws:
        raise ValueError(
            "This doesn't look like a Participant Breakdown — no "
            "'Participant'/'Participant $' header row was found."
        )

    def col(*names):
        for n in names:
            c = cols.get(_norm(n))
            if c:
                return c
        return None

    c_name = col("Participant")
    c_amt = col("Participant $")
    c_pct = col("Participant %")
    c_fee_pct = col("Origination %", "Points %")
    c_fee_amt = col("Origination $", "Points $")
    c_rate = col("Interest Rate")
    c_late = col("Share of Late fee", "Share of Late Fee")

    borrower = _value_right_of(ws, "Name")
    loan_amount = _num(_value_right_of(ws, "Loan amount"))
    loan_number = _value_right_of(ws, "Loan #")

    emails = _email_map(wb)

    def cell(r, c):
        return ws.cell(r, c) if c else None

    def frac_of(r, c, fallback):
        cl = cell(r, c)
        if cl is not None and cl.value is not None:
            return _to_fraction(cl.value)
        return fallback

    def pct_str(r, c, frac):
        cl = cell(r, c)
        d = _pct_decimals(cl.number_format if cl is not None else "")
        return f"{frac * 100:.{d}f}%"

    participants = []
    r = hdr_row + 1
    while c_name and r <= ws.max_row:
        name = ws.cell(r, c_name).value
        if name is None or not str(name).strip():
            break
        amount = _num(cell(r, c_amt).value) if cell(r, c_amt) else 0.0
        if amount <= 0:                       # skip $0 / non-participating rows
            r += 1
            continue

        part_f = frac_of(r, c_pct, amount / loan_amount if loan_amount else 0.0)
        fee_f = frac_of(r, c_fee_pct, 0.0)
        rate_f = frac_of(r, c_rate, 0.0)
        late_f = frac_of(r, c_late, part_f * 0.5)

        fee_amt_cell = cell(r, c_fee_amt)
        if fee_amt_cell is not None and fee_amt_cell.value is not None:
            fee_amt = _num(fee_amt_cell.value)
        else:
            fee_amt = amount * fee_f

        participants.append({
            "name": str(name).strip(),
            "amount": _money(amount),
            "participation_pct": pct_str(r, c_pct, part_f),
            "points_pct": pct_str(r, c_fee_pct, fee_f),
            "points_amount": _money(fee_amt),
            "interest_rate": pct_str(r, c_rate, rate_f),
            "late_fee_share_pct": pct_str(r, c_late, late_f),
            "email": emails.get(_norm(name), ""),
        })
        r += 1

    return {
        "deal": {
            "borrower_name": str(borrower).strip() if borrower else "",
            "loan_number": str(loan_number).strip() if loan_number is not None else "",
            "loan_amount": loan_amount,
        },
        "participants": participants,
    }
